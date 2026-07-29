import logging
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

BASE = Path(__file__).resolve().parent
CSV_DIR = BASE / "outputs" / "forecasts"
XLSX_DIR = BASE / "outputs" / "xlsx"

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALT_EVEN = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ALT_ODD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BLUE_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"), right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"), bottom=Side(style="thin", color="B4C6E7"),
)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F3864")
SUB_FONT = Font(name="Calibri", size=10, color="808080")


TURBINES = [f"TB{i:02d}" for i in range(1, 13)]


def _write_fast(df, xlsx_path):
    """Write DataFrame to xlsx using pandas engine (fast)."""
    df.to_excel(str(xlsx_path), index=False, engine="openpyxl")


def _write_power_forecast_by_turbine(csv_path, xlsx_path):
    """Split power_forecast.csv (5.6M rows) into per-turbine sheets."""
    logger.info("    Reading power_forecast.csv (5.6M rows)...")
    df = pd.read_csv(csv_path)
    total_rows = len(df)

    logger.info("    Writing per-turbine sheets...")
    with pd.ExcelWriter(str(xlsx_path), engine="openpyxl") as writer:
        # Summary sheet
        summary = pd.DataFrame({
            "Info": ["Power Forecast - AMG Wind Farm"],
            "Total Rows": [f"{total_rows:,}"],
            "Columns": ["timestamp_issue, timestamp_target, turbine_id, horizon_min, y_pred, y_low, y_high, model_version"],
        })
        summary.to_excel(writer, sheet_name="Summary", index=False)

        for tb in TURBINES:
            tb_df = df[df["turbine_id"] == tb].drop(columns=["turbine_id"], errors="ignore")
            if not tb_df.empty:
                tb_df.to_excel(writer, sheet_name=tb, index=False)

    return total_rows


def _open_and_format(xlsx_path, title, num_fmts=None, color_cols=None, color_map=None):
    """Re-open xlsx and apply styling (fast — no data writing)."""
    wb = load_workbook(str(xlsx_path))
    ws = wb.active

    ncols = ws.max_column
    nrows = ws.max_row

    # Insert title rows
    ws.insert_rows(1, 2)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tc = ws.cell(row=1, column=1, value=title)
    tc.font = TITLE_FONT
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    sc = ws.cell(row=2, column=1, value=f"{nrows - 1:,} rows | AMG Wind Farm Forecasting")
    sc.font = SUB_FONT
    sc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    # Style header (row 3 = first data header after title insert)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=3, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[3].height = 30
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{nrows + 1}"

    # Alternating row colors + borders
    for row in range(4, nrows + 3):
        fill = ALT_EVEN if row % 2 == 0 else ALT_ODD
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.fill = fill

    # Number formatting
    if num_fmts:
        for col_idx, fmt in num_fmts.items():
            for row in range(4, nrows + 3):
                cell = ws.cell(row=row, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = fmt

    # Conditional coloring on specific columns
    if color_cols:
        for col_idx, color_rule in color_cols.items():
            for row in range(4, nrows + 3):
                cell = ws.cell(row=row, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    if color_rule == "r2":
                        cell.fill = GREEN_FILL if cell.value >= 0.9 else (YELLOW_FILL if cell.value >= 0.7 else RED_FILL)
                    elif color_rule == "missing":
                        cell.fill = GREEN_FILL if cell.value == 0 else (YELLOW_FILL if cell.value < 10 else RED_FILL)
                    elif color_rule == "failure":
                        cell.fill = RED_FILL if cell.value >= 0.6 else (YELLOW_FILL if cell.value >= 0.4 else GREEN_FILL)
                    elif color_rule == "ci":
                        cell.fill = BLUE_FILL

    # Auto-width
    for col in range(1, ncols + 1):
        max_len = 8
        for row in range(3, min(nrows + 3, 200)):  # sample first 200 rows
            val = ws.cell(row=row, column=col).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)) + 2, 40))
        ws.column_dimensions[get_column_letter(col)].width = max_len

    wb.save(str(xlsx_path))


def convert_all():
    XLSX_DIR.mkdir(parents=True, exist_ok=True)
    csv_files = sorted(CSV_DIR.glob("*.csv"))
    results = []

    for csv_path in csv_files:
        xlsx_path = XLSX_DIR / f"{csv_path.stem}.xlsx"

        if csv_path.stat().st_size < 10:
            # Empty CSV — create minimal xlsx
            df = pd.DataFrame(columns=["Info"])
            df.loc[0] = ["No data available"]
            _write_fast(df, xlsx_path)
            results.append((csv_path.stem, 0))
            logger.info(f"  {csv_path.stem}.xlsx (empty)")
            continue

        df = pd.read_csv(csv_path)

        # Apply formatting based on file type
        name = csv_path.stem

        if name == "power_forecast":
            # Special handling: too large for single sheet, split by turbine
            total_rows = _write_power_forecast_by_turbine(csv_path, xlsx_path)
            results.append((name, total_rows))
            logger.info(f"  {name}.xlsx ({total_rows:,} rows, split by turbine)")
            continue

        _write_fast(df, xlsx_path)
        if name in ("metrics", "evaluation_metrics"):
            col_map = {c: i + 1 for i, c in enumerate(df.columns)}
            num_fmts = {}
            color_cols = {}
            for c in ["mae", "rmse", "MAE", "RMSE", "max_error"]:
                if c in col_map:
                    num_fmts[col_map[c]] = "#,##0.00"
            for c in ["nrmse_pct", "nRMSE", "r2", "R2", "skill_score"]:
                if c in col_map:
                    num_fmts[col_map[c]] = "0.0000"
            for c in ["n_samples"]:
                if c in col_map:
                    num_fmts[col_map[c]] = "#,##0"
            r2_key = col_map.get("R2", col_map.get("r2"))
            if r2_key:
                color_cols[r2_key] = "r2"
            _open_and_format(xlsx_path, f"{'Model Performance' if name == 'metrics' else 'Detailed Evaluation'} Metrics",
                             num_fmts=num_fmts, color_cols=color_cols)

        elif name == "farm_forecast":
            col_map = {c: i + 1 for i, c in enumerate(df.columns)}
            num_fmts = {}
            for c in ["farm_power_pred", "farm_energy_pred"]:
                if c in col_map:
                    num_fmts[col_map[c]] = "#,##0.00"
            _open_and_format(xlsx_path, "Farm-Level Forecast", num_fmts=num_fmts)

        elif name == "data_quality_report":
            col_map = {c: i + 1 for i, c in enumerate(df.columns)}
            color_cols = {}
            if "missing_rate" in col_map:
                color_cols[col_map["missing_rate"]] = "missing"
            num_fmts = {}
            if "missing_rate" in col_map:
                num_fmts[col_map["missing_rate"]] = "0.00"
            _open_and_format(xlsx_path, "Data Quality Report",
                             num_fmts=num_fmts, color_cols=color_cols)

        elif name == "ramp_alert":
            _open_and_format(xlsx_path, "Ramp Events Detected")

        elif name == "failure_risk":
            col_map = {c: i + 1 for i, c in enumerate(df.columns)}
            color_cols = {}
            num_fmts = {}
            if "failure_probability" in col_map:
                color_cols[col_map["failure_probability"]] = "failure"
                num_fmts[col_map["failure_probability"]] = "0.0000"
            _open_and_format(xlsx_path, "Turbine Failure Risk",
                             num_fmts=num_fmts, color_cols=color_cols)

        elif name == "anomaly_alert":
            _open_and_format(xlsx_path, "Anomaly Alerts")

        elif name == "temperature_warning":
            col_map = {c: i + 1 for i, c in enumerate(df.columns)}
            color_cols = {}
            num_fmts = {}
            if "temperature" in col_map:
                num_fmts[col_map["temperature"]] = "0.00"
            _open_and_format(xlsx_path, "Temperature Warnings",
                             num_fmts=num_fmts, color_cols=color_cols)

        elif name == "forecasts":
            _open_and_format(xlsx_path, "Forecast Output")

        else:
            _open_and_format(xlsx_path, name.replace("_", " ").title())

        results.append((name, len(df)))
        logger.info(f"  {name}.xlsx ({len(df):,} rows)")

    return results


if __name__ == "__main__":
    logger.info("=" * 50)
    logger.info("Converting CSV -> XLSX (fast mode)")
    logger.info("=" * 50)

    results = convert_all()

    logger.info("=" * 50)
    logger.info(f"Converted {len(results)} files")
    for name, rows in results:
        logger.info(f"  {name}: {rows:,} rows")
    logger.info(f"Output: {XLSX_DIR}")
    logger.info("=" * 50)
